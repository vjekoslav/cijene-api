import datetime
import json
import logging
import re
from io import BytesIO
from tempfile import TemporaryFile
from typing import Any, List

import openpyxl
from crawler.store.models import Product, Store

from .base import BaseCrawler

logger = logging.getLogger(__name__)


class DmCrawler(BaseCrawler):
    """Crawler for DM (DrogerieMarkt) store prices."""

    CHAIN = "dm"
    BASE_URL = "https://www.dm.hr"
    CONTENT_BASE_URL = "https://content.services.dmtech.com/rootpage-dm-shop-hr-hr"
    INDEX_URL = f"{CONTENT_BASE_URL}/novo/promocije/nove-oznake-cijena-i-vazeci-cjenik-u-dm-u-2906632?mrclx=false"

    # DM has global prices, not per-store prices
    STORE_ID = "all"
    STORE_NAME = "DM"

    def parse_date_from_title(self, title: str) -> datetime.date:
        """
        Extract date from the title the Excel link.

        Args:
            title: Title attribute of the link

        Returns:
            Extracted date object
        """
        # Match date in format DD.MM.YYYY where D or M can be single-digit
        date_match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", title)
        if not date_match:
            raise ValueError(f"Could not extract date from title: {title}")

        day, month, year = map(int, date_match.groups())
        return datetime.date(year, month, day)

    def find_excel_url(self, json_content: str, target_date: datetime.date) -> str:
        """
        Parse the JSON data to find the Excel file URL for the target date.

        Args:
            json_content: JSON content from the index page
            target_date: The date to search for

        Returns:
            URL of the Excel file

        Raises:
            ValueError: If no Excel file is found for the target date
        """
        try:
            # Parse JSON data
            data = json.loads(json_content)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            raise ValueError("Failed to parse JSON data from the page")

        # Find all CMDownload entries in mainData
        excel_entries = []
        for item in data.get("mainData", []):
            if item.get("type") == "CMDownload":
                excel_entries.append(item.get("data", {}))

        if not excel_entries:
            logger.warning("No Excel links found in JSON data")
            raise ValueError("No Excel links found in JSON data")

        target_date_str = f"{target_date.day}.{target_date.month}.{target_date.year}"
        logger.info(f"Looking for Excel file with date {target_date_str}")

        for entry in excel_entries:
            headline = entry.get("headline", "")
            link_target = entry.get("linkTarget", "")

            if not headline or not link_target:
                continue

            try:
                link_date = self.parse_date_from_title(headline)
                if link_date == target_date:
                    # Ensure URL is absolute
                    if not link_target.startswith(("http://", "https://")):
                        url = f"{self.CONTENT_BASE_URL}{link_target}"
                    else:
                        url = link_target
                    logger.info(f"Found Excel file with date {link_date}: {url}")
                    return url
            except Exception as e:
                logger.warning(f"Error parsing date from headline '{headline}': {e}")
                continue

        raise ValueError(f"No Excel file found for date {target_date_str}")

    def detect_columns(self, worksheet: Any) -> list[str]:
        """
        Detect the column ordering in the DM Excel worksheet.

        This relies on the fact that one of the columns in the header will
        always be "naziv + šifra", which is a merged cell that actually
        has two cells in the data, naziv and product ID.

        Args:
            worksheet: The active worksheet object

        Returns:
            List of column headers
        """

        def fix_col_name(name: str | None) -> str:
            """
            Normalize column names.
            """
            name = str(name or "")
            name = self.strip_diacritics(name.lower())
            words = name.split()
            return " ".join(w for w in words if w)
            return self.strip_diacritics(name.lower().replace("\t", " "))

        for row in worksheet.iter_rows():
            row_str = [fix_col_name(cell.value) for cell in row]
            if "naziv + sifra" in row_str:
                idx = row_str.index("naziv + sifra")
                if row_str[idx + 1] != "":
                    raise ValueError(
                        "Expected 'naziv + šifra' to be a merged cell with two parts"
                    )
                row_str[idx] = "naziv"
                row_str[idx + 1] = "sifra"
                return row_str

        raise ValueError(
            "Could not detect Excel columns, DM file format may have changed"
        )

    @staticmethod
    def map_columns(row: Any, columns: list) -> dict[str, Any]:
        """
        Map the row data to a dictionary using the detected columns.

        Args:
            row: The row object from the worksheet
            columns: List of column headers

        Returns:
            Dictionary mapping column names to cell values
        """
        return {col: str(row[i].value or "").strip() for i, col in enumerate(columns)}

    def parse_excel(self, excel_data: bytes) -> List[Product]:
        """
        Parse Excel file data into Product objects.

        Args:
            excel_data: Raw Excel file content

        Returns:
            List of Product objects
        """
        logger.debug("Parsing Excel file")
        products = []

        try:
            workbook = openpyxl.load_workbook(BytesIO(excel_data), data_only=True)
            worksheet = workbook.active  # Get the active worksheet

            if not worksheet:
                raise ValueError("No active worksheet found in the Excel file")

            columns = self.detect_columns(worksheet)
            logger.debug(f"Detected columns: {columns}")

            for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
                # Skip header and empty rows
                if len(row) != len(columns):
                    continue

                row_map = self.map_columns(row, columns)
                if not row_map["sifra"]:
                    continue

                try:
                    product_data = {
                        "product": row_map["naziv"],
                        "product_id": row_map["sifra"],
                        "brand": row_map["marka"],
                        "barcode": row_map["barkod"],
                        "category": row_map["kategorija proizvoda"],
                        "quantity": row_map["neto kolicina"],
                        "unit": row_map["jedinica mjere"],
                        "unit_price": self.parse_price(
                            row_map["cijena za jedinicu mjere"], False
                        ),
                        "price": self.parse_price(row_map["mpc"], False),
                        "special_price": self.parse_price(
                            row_map[
                                "mpc za vrijeme posebnog oblika prodaje (rasprodaja proizvoda koji izlaze iz asortimana)"
                            ],
                            False,
                        ),
                        "best_price_30": self.parse_price(
                            row_map[
                                "najniza cijena u posljednjih 30 dana prije rasprodaje"
                            ],
                            False,
                        ),
                        "anchor_price": self.parse_price(
                            row_map[
                                "sidrena cijena na 2.5.2025. ili na datum ulistanja"
                            ],
                            False,
                        ),
                    }

                    # Apply common fixups from base class
                    product_data = self.fix_product_data(product_data)

                    # Create Product object
                    product = Product(**product_data)  # type: ignore
                    products.append(product)
                except Exception as e:
                    row_txt = "; ".join([str(cell.value or "") for cell in row])
                    logger.warning(f"Failed to parse row {row_idx}: `{row_txt}`: {e}")
                    continue

        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            raise

        logger.debug(f"Parsed {len(products)} products from Excel file")
        return products

    def get_all_products(self, date: datetime.date) -> list[Store]:
        """
        Main method to fetch and parse all product and price info.

        Args:
            date: The date to search for in the price list.

        Returns:
            List with a single Store object containing all products.

        Raises:
            ValueError: If no price list is found for the given date.
        """
        content = self.fetch_text(self.INDEX_URL)
        if not content:
            logger.warning(f"No content found at {self.INDEX_URL}")
            return []

        # Find Excel file URL for the exact target date from JSON
        excel_url = self.find_excel_url(content, date)
        logger.info(f"Found Excel file URL: {excel_url}")

        # Download Excel file
        with TemporaryFile(mode="w+b") as temp_file:
            self.fetch_binary(excel_url, temp_file)
            temp_file.seek(0)
            excel_data = temp_file.read()

        # Parse Excel file
        products = self.parse_excel(excel_data)

        if not products:
            logger.warning(f"No products found for date {date}")
            return []

        # Create a global store
        store = Store(
            chain=self.CHAIN,
            store_type="store",
            store_id=self.STORE_ID,
            name=self.STORE_NAME,
            street_address="",
            zipcode="",
            city="",
            items=products,
        )

        return [store]


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    crawler = DmCrawler()
    stores = crawler.crawl(datetime.date.today())
    print(stores[0])
    print(stores[0].items[0])
